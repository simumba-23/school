from django.shortcuts import render
from openpyxl import load_workbook
from .models import Product

def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, price, quantity = row
            Product.objects.create(name=name, price=price, quantity=quantity)

        return render(request, 'import_success.html')

    return render(request, 'import_form.html')


from django.http import HttpResponse
from openpyxl import Workbook
from .models import Product



def export_to_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="products.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Add headers
    headers = ["Name", "Price", "Quantity"]
    ws.append(headers)

    # Add data from the model
    products = Product.objects.all()
    for product in products:
        ws.append([product.name, product.price, product.quantity])

    # Save the workbook to the HttpResponse
    wb.save(response)
    return response